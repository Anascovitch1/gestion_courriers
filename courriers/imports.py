"""Import de la base des contribuables depuis un fichier Excel (.xlsx)."""
from openpyxl import load_workbook

from .models import Contribuable

# Correspondance des en-têtes acceptés -> champ du modèle
COLONNES = {
    "nif": "nif",
    "raison sociale": "raison_sociale",
    "raison_sociale": "raison_sociale",
    "activite": "activite",
    "activité": "activite",
    "adresse": "adresse",
    "telephone": "telephone",
    "téléphone": "telephone",
    "email": "email",
    "e-mail": "email",
    "regime": "regime",
    "régime": "regime",
    "regime fiscal": "regime",
}


def importer_contribuables(fichier):
    """Lit le fichier et crée/met à jour les contribuables. Retourne (créés, maj, erreurs)."""
    wb = load_workbook(fichier, read_only=True, data_only=True)
    ws = wb.active

    lignes = ws.iter_rows(values_only=True)
    entetes = next(lignes, None)
    if not entetes:
        return 0, 0, ["Fichier vide."]

    # index de colonne -> champ
    mapping = {}
    for i, libelle in enumerate(entetes):
        if libelle is None:
            continue
        cle = str(libelle).strip().lower()
        if cle in COLONNES:
            mapping[i] = COLONNES[cle]

    if "nif" not in mapping.values():
        return 0, 0, ["Colonne « NIF » introuvable dans le fichier."]

    crees = maj = 0
    erreurs = []
    for num, ligne in enumerate(lignes, start=2):
        donnees = {champ: (str(ligne[i]).strip() if ligne[i] is not None else "")
                   for i, champ in mapping.items()}
        nif = donnees.get("nif", "").strip()
        if not nif:
            continue
        donnees.pop("nif")
        try:
            obj, cree = Contribuable.objects.update_or_create(nif=nif, defaults=donnees)
            crees += 1 if cree else 0
            maj += 0 if cree else 1
        except Exception as exc:  # noqa: BLE001
            erreurs.append(f"Ligne {num} (NIF {nif}) : {exc}")
    wb.close()
    return crees, maj, erreurs
